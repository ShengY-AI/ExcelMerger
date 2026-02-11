import os
import sys
import csv
import re
from datetime import datetime
from collections import OrderedDict
from openpyxl import Workbook


# =========================
# 基础路径/文件夹
# =========================

def app_dir() -> str:
    """返回运行基准目录：
    - 开发态：.py 所在目录
    - 打包态 Windows：exe 所在目录
    - 打包态 macOS：.app 的父目录（即与 InputFolder/OutputFolder 同级的发布目录）
    """
    if getattr(sys, "frozen", False):
        exe_path = sys.executable.replace("\\", "/")
        marker = ".app/Contents/MacOS"
        if marker in exe_path:
            app_path = exe_path.split(marker)[0] + ".app"
            return os.path.dirname(app_path)  # ✅ 关键：返回 .app 的父目录
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def ensure_folders(base: str):
    input_dir = os.path.join(base, "InputFolder")
    output_dir = os.path.join(base, "OutputFolder")
    os.makedirs(input_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)
    return input_dir, output_dir


# =========================
# 自然排序（2 < 10）
# =========================

def natural_key(text: str):
    return [int(t) if t.isdigit() else t.lower()
            for t in re.split(r"(\d+)", text)]


# =========================
# position 标准化（关键修复）
# =========================

def normalize_position(pos: str) -> str:
    """
    规范化 position：
    - 去空格
    - 大小写统一
    - 若包含 '+'：拆分后排序再拼接，使 opl+ipl 与 ipl+opl 归一
    """
    if pos is None:
        return ""
    pos = pos.strip().lower().replace(" ", "")
    if "+" in pos:
        parts = [p for p in pos.split("+") if p]
        parts.sort()  # 使 ipl+opl 与 opl+ipl 得到同一表示
        pos = "+".join(parts)
    return pos


# =========================
# 命名解析（文件夹名 / CSV名）
# 支持：
# - ..._0004(microglia-opl)_Detailed.csv / _Statistics.csv
# - ... (microglia-opl)_Statistics  (无编号)
# =========================

NAME_RE = re.compile(
    r"""
    ^(?P<mouse>p\d+)
    -.*?
    (?:_(?P<view>\d{3,6}))?
    \(
      (?P<kind>microglia|ribeyes)
      -
      (?P<pos>[^)]+)
    \)
    _(?P<tag>Detailed|Statistics)?
    (?:\.csv)?$
    """,
    re.IGNORECASE | re.VERBOSE
)


def parse_from_name(name: str):
    """
    返回 (mouse, view, pos, kind)；解析失败返回 None
    view 缺失 -> "/"
    pos 做 normalize_position
    """
    m = NAME_RE.match(name)
    if not m:
        return None

    mouse = m.group("mouse").lower()
    view = m.group("view") or "/"
    pos = normalize_position(m.group("pos") or "")
    kind = (m.group("kind") or "").lower()
    return mouse, view, pos, kind


# =========================
# 读取 CSV：提取第一列第一个 float
# =========================

def first_float_in_first_column(csv_path: str):
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            if not row:
                continue
            s = str(row[0]).strip()
            if not s:
                continue
            try:
                return float(s)
            except ValueError:
                continue
    return None


# =========================
# 扫描：以“子文件夹”为单位合并
# =========================

def list_all_dirs(input_dir: str):
    dirs = []
    for root, subdirs, files in os.walk(input_dir):
        subdirs[:] = [d for d in subdirs if not d.startswith(".")]
        if root != input_dir:
            dirs.append(root)
    dirs.sort(key=lambda p: natural_key(os.path.relpath(p, input_dir)))
    return dirs


def list_csv_in_dir(dir_path: str):
    csvs = []
    for f in os.listdir(dir_path):
        if f.startswith("~$"):
            continue
        if f.lower().endswith(".csv"):
            csvs.append(os.path.join(dir_path, f))
    csvs.sort(key=lambda p: natural_key(os.path.basename(p)))
    return csvs


def build_summary(input_dir: str):
    summary = OrderedDict()
    all_dirs = list_all_dirs(input_dir)

    for d in all_dirs:
        folder_name = os.path.basename(d)

        # 1) 优先从文件夹名解析 mouse/view/pos（更可靠）
        folder_parsed = parse_from_name(folder_name)

        csvs = list_csv_in_dir(d)
        if not csvs:
            continue

        fixed_mouse = fixed_view = fixed_pos = None
        if folder_parsed is not None:
            fixed_mouse, fixed_view, fixed_pos, _ = folder_parsed

        # 2) 遍历 csv：决定 ribeyes / microglia 并写入
        for csv_path in csvs:
            csv_name = os.path.basename(csv_path)

            if fixed_mouse is not None:
                mouse, view, pos = fixed_mouse, fixed_view, fixed_pos
                parsed_csv = parse_from_name(csv_name)
                kind = parsed_csv[3] if parsed_csv else None
            else:
                parsed_csv = parse_from_name(csv_name)
                if parsed_csv is None:
                    continue
                mouse, view, pos, kind = parsed_csv

            # kind 兜底
            if not kind:
                lower = csv_name.lower()
                if "ribeyes" in lower:
                    kind = "ribeyes"
                elif "microglia" in lower:
                    kind = "microglia"
                else:
                    continue

            val = first_float_in_first_column(csv_path)

            # ✅ key 使用规范化后的 pos（opl+ipl 与 ipl+opl 会合并）
            key = (mouse, view, pos)
            if key not in summary:
                summary[key] = {
                    "mouse ID": mouse,
                    "view-id": view,
                    "position": pos,
                    "Ribeyes/μm": None,
                    "microglia/μm": None,
                    "engulfment index": None,
                }

            if kind == "ribeyes":
                summary[key]["Ribeyes/μm"] = val
            elif kind == "microglia":
                summary[key]["microglia/μm"] = val

    return summary


def write_excel(summary: OrderedDict, output_dir: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    headers = ["mouse ID", "view-id", "position", "Ribeyes/μm", "microglia/μm", "engulfment index"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c).value = h

    r = 2
    for _, row in summary.items():
        ws.cell(row=r, column=1).value = row["mouse ID"]
        ws.cell(row=r, column=2).value = row["view-id"]
        ws.cell(row=r, column=3).value = row["position"]
        ws.cell(row=r, column=4).value = row["Ribeyes/μm"]
        ws.cell(row=r, column=5).value = row["microglia/μm"]
        ws.cell(row=r, column=6).value = row["engulfment index"]
        r += 1

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(output_dir, f"merged_summary_{ts}.xlsx")
    wb.save(out_path)
    return out_path


def main():
    base = app_dir()
    input_dir, output_dir = ensure_folders(base)

    summary = build_summary(input_dir)
    if not summary:
        print("未提取到数据：请确认 InputFolder 子目录中有 .csv，且目录名/文件名含 (ribeyes-XXX) 或 (microglia-XXX)。")
        return

    out = write_excel(summary, output_dir)
    print(f"完成 ✅ 输出文件：{out}")


if __name__ == "__main__":
    main()
